import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

filePath = "C:/Users/89704/OneDrive/Documents/我的文档/工作/RPA测试文档/数据库设计说明书-表结构设计.xlsx"
sheetName = "管理员维护表"

workbook = openpyxl.load_workbook(filePath)
worksheet = workbook[sheetName]


fill = PatternFill(fgColor="808080")
# 设置列头单元格样式
for cell in worksheet[1]:
    cell.fill = fill

# 保存Excel文件
workbook.save(filePath)