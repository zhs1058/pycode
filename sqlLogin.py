import threading
import tkinter as tk
from tkinter import filedialog
from tkinter import *
import psycopg2
from psycopg2 import extensions
import openpyxl
import requests, json, datetime


class LoginDialog:

    def __init__(self, title):
        self.root = tk.Tk()
        self.root.resizable(0, 0)
        self.root.attributes('-topmost', True)
        self.url = ''
        self.port = ''
        self.dataBase = ''
        self.schema = ''
        self.user = ''
        self.pwd = ''
        self.documentDir = ''
        # icoPath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "psbc1.ico")
        # self.root.iconbitmap(icoPath)
        self.root.title(title)
        # 获取屏幕大小
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        windows_width = round(950 / 2240 * screen_width)
        windows_height = round(600 / 1400 * screen_height)
        local_width = screen_width // 3
        local_height = screen_height // 3
        self.root.geometry(f"{windows_width}x{windows_height}+{local_width}+{local_height}")
        self.root.tips = StringVar()
        self.timerTips = threading.Timer(0.5, lambda: self.root.tips.set(''))

        # url输入框
        entry_width = round(17 / windows_width * 650 * 1.0)
        tk.Label(self.root, text="地址：", font=("宋体", 20)).place(relx=0.05, rely=0.05)
        self.url_entry = tk.Entry(self.root, font=("宋体", 20), width=f"{entry_width}")
        self.url_entry.place(relx=0.22, rely=0.05)

        # 端口号输入框
        entry_width = round(17 / windows_width * 650 * 0.3)
        tk.Label(self.root, text="端口：", font=("宋体", 20)).place(relx=0.65, rely=0.05)
        self.port_entry = tk.Entry(self.root, font=("宋体", 20), width=f"{entry_width}")
        self.port_entry.place(relx=0.8, rely=0.05)
        self.port_entry.insert(0, "5432")

        # 数据库输入框
        entry_width = round(17 / windows_width * 650 * 1.0)
        tk.Label(self.root, text="数据库：", font=("宋体", 20)).place(relx=0.05, rely=0.17)
        self.data_base_entry = tk.Entry(self.root, font=("宋体", 20), width=f"{entry_width}")
        self.data_base_entry.place(relx=0.22, rely=0.17)

        # 模式输入框
        entry_width = round(17 / windows_width * 650 * 1.0)
        tk.Label(self.root, text="模式：", font=("宋体", 20)).place(relx=0.05, rely=0.29)
        self.schema_entry = tk.Entry(self.root, font=("宋体", 20), width=f"{entry_width}")
        self.schema_entry.place(relx=0.22, rely=0.29)
        # 用户名输入框
        entry_width = round(17 / windows_width * 650 * 1.0)
        tk.Label(self.root, text="用户名：", font=("宋体", 20)).place(relx=0.05, rely=0.41)
        self.user_entry = tk.Entry(self.root, font=("宋体", 20), width=f"{entry_width}")
        self.user_entry.place(relx=0.22, rely=0.41)

        # 密码输入框
        entry_width = round(17 / windows_width * 650 * 1.0)
        tk.Label(self.root, text="密码：", font=("宋体", 20)).place(relx=0.05, rely=0.53)
        self.pwd_entry = tk.Entry(self.root, font=("宋体", 20), width=f"{entry_width}")
        self.pwd_entry.place(relx=0.22, rely=0.53)

        # 选择文件
        self.document_button = tk.Button(self.root, text="表结构设计文件", width="15", font=("宋体", 15),
                                         command=self.select_file)
        self.document_button.place(relx=0.65, rely=0.53)

        # 提示信息
        self.root.message = tk.Message(self.root, width="600", textvariable=self.root.tips, fg="red", font=("宋体", 15))
        self.root.message.place(relx=0.02, rely=0.63)

        #
        # 确定和取消按钮
        self.ok_button = tk.Button(self.root, text="确定", width="10", font=("宋体", 15), command=self.on_ok)
        self.ok_button.place(relx=0.3, rely=0.83)
        self.cancel_button = tk.Button(self.root, text="取消", width="10", font=("宋体", 15), command=self.on_cancel)
        self.cancel_button.place(relx=0.6, rely=0.83)

        # 登录结果
        self.result = None
        # 是否成功获取登录结果标识
        self.resultFlag = False
        # 结果key值列
        self.resultKeyList = []
        # 结果字典
        self.resultDict = {}

    def showTips(self, msg):
        self.root.tips.set(msg)
        self.timerTips = threading.Timer(5.0, lambda: self.root.tips.set(''))
        self.timerTips.start()

    def select_file(self):
        file_path = filedialog.askopenfilename()
        self.documentDir = file_path
        self.showTips("请注意：所选择的文件的Sheet页<系统表清单>中的“工程名称”，“表英文名称”，“表英文名称”三列必须填写完整，否则无法生成各个表详细信息")

    def on_ok(self):
        # 检查地址
        if self.url_entry.get() == '':
            self.showTips('请输入地址！')
            return
        self.url = self.url_entry.get()
        # 检查端口
        if self.port_entry.get() == '':
            self.showTips('请输入端口号！')
            return
        self.port = self.port_entry.get()
        # 检查数据库
        if self.data_base_entry.get() == '':
            self.showTips('请输入数据库！')
            return
        self.dataBase = self.data_base_entry.get()
        # 检查模式
        if self.schema_entry.get() == '':
            self.showTips('请输入模式！')
            return
        self.schema = self.schema_entry.get()
        # 检查用户名
        if self.user_entry.get() == '':
            self.showTips('请输入用户名！')
            return
        self.user = self.user_entry.get()
        # 检查文件
        if self.documentDir == '':
            self.showTips('请选择文件！')
            return

        # 访问数据库
        self.getData()
        # 后续处理
        self.resultFlag = True
        if self.timerTips.is_alive():
            self.timerTips.cancel()
        self.root.destroy()

    def on_cancel(self):
        self.result = "取消"  # 取消登录时设置结果为字符串"取消登录"
        if self.timerTips.is_alive():
            self.timerTips.cancel()
        self.root.destroy()

    def show(self):
        self.root.mainloop()

    def openExcel(self):
        # 指定要打开的Excel文件路径
        excel_path = self.documentDir

        # 打开Excel文件
        workbook = openpyxl.load_workbook(excel_path)

        # 选择要搜索的工作表
        sheet_name = "系统表清单"
        sheet = workbook[sheet_name]

        # 列名所在的行数
        column_names_row = 1

        # 获取列名所对应的列索引
        column_names = {
            "工程名称（必填）": None,
            "英文表名（必填）": None,
            "中文表名（必填）": None
        }

        for column in sheet.iter_cols():
            cell_value = column[0].value
            if cell_value in column_names:
                column_names[cell_value] = column[0].column_letter
        my_list = []
        # 循环打印指定列的内容
        for row in sheet.iter_rows(min_row=column_names_row + 1):
            project_name = sheet[column_names["工程名称（必填）"] + str(row[0].row)].value
            english_name = sheet[column_names["英文表名（必填）"] + str(row[0].row)].value
            chinese_name = sheet[column_names["中文表名（必填）"] + str(row[0].row)].value
            if not str(english_name).startswith("英文"):
                pair = {"project": project_name, "english": english_name, "chinese": chinese_name}
                my_list.append(pair)
            # print(f"工程名称: {project_name}")
            # print(f"英文表名: {english_name}")
            # print(f"中文表名: {chinese_name}")
            # print("-----")

        # 关闭Excel文件
        workbook.close()
        return my_list

    def getData(self):
        # 连接到数据库
        try:
            conn = psycopg2.connect(
                host="120.53.224.82",
                port=5432,
                database="ibank",
                user="ibank",
                password="UHNiY0NvcnB3bEAyMDIx"
            )
        except Exception as e:
            self.showTips(str(e).replace("\n", ""))

        # 设置连接的事务隔离级别为自动提交
        conn.set_isolation_level(extensions.ISOLATION_LEVEL_AUTOCOMMIT)

        # 创建一个游标对象
        cur = conn.cursor()

        # 获取表的详细结构信息
        def get_table_structure(table_name, project, chinese):
            # 查询列信息
            cur.execute(
                f"""
                select 
                    col.ordinal_position as serial, 
                    col.table_name as tableName, 
                    col.column_name as columnName,
                    col.udt_name as dataType, 
                    COALESCE(col.character_maximum_length, col.numeric_precision, col.datetime_precision) as len,
                    col.numeric_scale as acc,
                    col.is_nullable as isNull,
                    col.column_default as defaultValue,
                    des.description as desc,
                    def.def_forkey as foreignkey
                from
        	        information_schema.columns col 
        	        left join pg_description des on col.table_name::regclass = des.objoid and col.ordinal_position = des.objsubid
        	        left join 
        		    (SELECT
        			    kcu.column_name as def_column, 
        			    ccu.table_name || '.' || ccu.column_name AS def_forkey
        		    FROM 
                        information_schema.table_constraints AS tc 
                        JOIN information_schema.key_column_usage AS kcu ON tc.constraint_name = kcu.constraint_name
                        JOIN information_schema.constraint_column_usage AS ccu ON ccu.constraint_name = tc.constraint_name
        		    WHERE constraint_type = 'FOREIGN KEY' AND tc.table_name = '{table_name}') def on def.def_column = col.column_name
                where
        	        table_name = '{table_name}'
                order by 
        	        ordinal_position;""")
            columns = cur.fetchall()

            # 打印表的详细结构信息
            dataList = []

            for column in columns:
                serial, tableName, columnName, dataType, len, acc, isNull, defaultValue, desc, foreignkey = column
                colList = []

                # 字段序号
                colList.append(serial)
                # 工程名称
                colList.append(project)
                # 英文表名
                colList.append(str(tableName).upper())
                # 中文表名
                colList.append(chinese)
                # 字段英文名
                colList.append(str(columnName).upper())
                # 字段中文名
                colList.append(desc)
                # 字段类型
                colList.append(str(dataType).upper())
                # 长度
                if len is None:
                    len = 0
                colList.append(len)
                # 精度
                if acc is None:
                    acc = 0
                colList.append(acc)
                # 是否主键
                pKey = "N"
                if serial == 1:
                    pKey = "Y"
                colList.append(pKey)
                # 外键
                if foreignkey is None:
                    foreignkey = "/"
                else:
                    foreignkey = str(foreignkey).upper()
                colList.append(foreignkey)
                # 是否可以为空
                if isNull == "YES":
                    isNull = "Y"
                else:
                    isNull = "N"
                colList.append(isNull)
                # 缺省值
                if defaultValue is None:
                    defaultValue = "/"
                colList.append(defaultValue)
                # 取值范围
                colList.append("/")
                # 业务取值说明
                colList.append("/")
                # 当前状态
                colList.append("无变化")
                # 字段含义
                colList.append(desc)
                # 是否贯标
                colList.append("N")
                # 企业级数据字典数据
                colList.append("/")
                # 不贯标原因
                colList.append("自定义字段")

                dataList.append(colList)

            return dataList

        # 获取表的汇总信息
        dataList = self.openExcel()
        for item in dataList:
            # 调用函数获取表结构信息
            tableList = get_table_structure(str(item["english"]).lower(), item["project"], item["chinese"])
            self.resultDict[item["chinese"]] = tableList
            self.resultKeyList.append(item["chinese"])

        # 关闭游标和连接
        cur.close()
        conn.close()


dialog = LoginDialog("连接数据库")
dialog.show()
if dialog.resultFlag:
    revar1 = dialog.resultKeyList
    revar2 = dialog.resultDict
    revar3 = dialog.documentDir
else:
    revar3 = 'N'

