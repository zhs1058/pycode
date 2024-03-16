import sys
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, numbers
import tkinter as tk
from tkinter import filedialog, ttk
from ttkbootstrap import Style


class honmeDialog:

    def __init__(self):
        self.root = tk.Tk()
        self.root.resizable(0, 0)
        self.root.attributes('-topmost', True)
        self.root.title("选择文件")
        self.root.geometry("600x500+500+100")
        # 增加注释
        style = Style(theme='litera')


        self.tip = tk.Label(self.root, wraplength="400", text="", fg="red", font=("宋体", 15))
        self.tip.place(x=50, y=10)
        self.msg = tk.Label(self.root, wraplength="400", text="", fg="green", font=("宋体", 15))
        self.msg.place(x=90, y=30)

        # 暂估应付款
        ttk.Label(self.root, text="暂估应付款：", font=("宋体", 15)).place(x=90, y=148)
        self.select_button = tk.Button(self.root, text="选择", width="10", font=("宋体", 15), command=self.on_select_1)
        self.select_button.place(x=330, y=147)

        # 明细应付款
        tk.Label(self.root, text="明细应付款：", font=("宋体", 15)).place(x=90, y=188)
        self.select_button = tk.Button(self.root, text="选择", width="10", font=("宋体", 15), command=self.on_select_2)
        self.select_button.place(x=330, y=187)

        # 生成文件路径
        tk.Label(self.root, text="生成文件路径：", font=("宋体", 15)).place(x=90, y=260)
        self.select_button = tk.Button(self.root, text="选择", width="10", font=("宋体", 15),
                                       command=self.on_select_dir)
        self.select_button.place(x=330, y=260)

        # 确认按钮
        self.ok_button = tk.Button(self.root, text="生成", width="10", font=("宋体", 15), command=self.on_ok)
        self.ok_button.place(x=330, y=350)

        # 结果
        self.filePath_1 = ""
        self.fileDir = ""

        self.resultMap = {}
        self.zgTempResultMap = {}
        self.mxTempResultMap = {}
        self.tempResultMap = {}
        self.result = []
        self.titleIndexMap = {9: "经济金融类", 10: "国际贸易融资类", 11: "财务审计类",
                              12: "风险管理类", 13: "法律类", 14: "计算机类",
                              15: "人力资源类", 16: "其他", 17: "职称类证书"}

    def on_ok(self):
        if (self.filePath_1 == ""):
            self.tip.config(text=f"请选择选择证书统计表")
            return
        if (self.fileDir == ""):
            self.tip.config(text=f"请选择生成文件目录")
            return
        self.tip.config(text=f"")
        self.msg.config(text=f"开始执行,请稍后...")

        self.fill_data()
        self.on_close()

    def on_cancel(self):
        self.on_close()

    def on_close(self):
        self.root.destroy()

    def show(self):
        self.root.mainloop()

    def on_select_1(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.tip.config(text=f"路径:{file_path}")
            self.filePath_1 = file_path
            self.do_excel_1(self.filePath_1)

    def on_select_2(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.tip.config(text=f"路径:{file_path}")
            self.filePath_1 = file_path
            self.do_excel(self.filePath_1)

    def on_select_dir(self):
        selected_path = filedialog.askdirectory()

        # 如果选择了路径，则打印路径
        if selected_path:
            self.tip.config(text=f"选择的路径:{selected_path}")
            self.fileDir = selected_path

    # -------------------------------------------------------------

    def fill_data(self):
        # 输出文件路径（新的Excel表格）
        output_file_path = self.fileDir + "/采购对账单.xlsx"

        # if key not in self.tempResultMap:
        #     self.tempResultMap[key] = row
        # else:
        #     preRow = self.tempResultMap.get(key, "default")
        #     if preRow[1] < row[1]:
        #         self.tempResultMap[key] = row


        for key, value in self.resultMap.items():
            if value[2] is None and value[3] is None and value[4] is not None:
                value[5] = value[4]
            if value[2] is None and value[3] is not None and value[4] is not None:
                value[5] = value[4] + value[3]
            if value[2] is not None and value[3] is None and value[4] is not None:
                temp = value[4] - value[2]
                if temp != 0:
                    value[5] = temp
                else:
                    value[5] = None
            if value[2] is not None and value[3] is None and value[4] is None:
                value[5] = 0 - value[2]
            if value[2] is None and value[3] is not None and value[4] is None:
                value[5] = value[3]
            self.result.append(value)


        self.set_excel_style(output_file_path)

    # 设置excel表格式
    def set_excel_style(self, output_file_path):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet("统计", index=0)

        self.set_excel_col_width(worksheet)
        self.set_excel_data(worksheet)

        workbook.save(output_file_path)

    def building_data(self, data):
        itemResult = []
        for key, value in self.titleIndexMap.items():
            if data[key] is not None:
                v = data[key]
                vs = v.split(',')
                for vItem in vs:
                    itemResult = [data[0], data[1], data[3], data[6], data[7], data[8], value, vItem]
                    self.result.append(itemResult)


    def do_excel_1(self, input_file_path):
        # 打开Excel文件
        wb_now = openpyxl.load_workbook(input_file_path, data_only=True)
        # 选择倒数第一个工作表
        sheet_now = wb_now.worksheets[-1]

        self.read_excel_1(sheet_now)
        # 关闭Excel文件
        wb_now.close()

    def do_excel(self, input_file_path):
        # 打开Excel文件
        wb_now = openpyxl.load_workbook(input_file_path, data_only=True)
        # 选择倒数第一个工作表
        sheet_now = wb_now.worksheets[-1]

        self.read_excel(sheet_now)
        # 关闭Excel文件
        wb_now.close()

    def read_excel_1(self, sheet):
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # 遍历每一行
            key = row[0]
            sj = 0
            if row[2] is not None:
                sj = round((row[2] * 1.13), 2)
            elif row[3] is not None:
                sj = round((row[3] * 1.13), 2)

            if key in self.resultMap:
                preResult = self.resultMap[key]
                preResult[4] = sj
                self.resultMap[key] = preResult
            else:
                itemResult = [row[0], row[1], None, None, sj, None]
                self.resultMap[key] = itemResult




    # 读取行数
    def read_excel(self, sheet):
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # 遍历每一行
            key = row[0]
            if key in self.resultMap:
                preResult = self.resultMap[key]
                preResult[2] = row[2]
                preResult[3] = row[3]
                self.resultMap[key] = preResult
            else:
                itemResult = [row[0], row[1], row[2], row[3], None, None]
                self.resultMap[key] = itemResult


    # 设置列宽
    def set_excel_col_width(self, worksheet):
        for i in range(1, 7):
            if i == 2:
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 40
            else:
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15

    def set_excel_data(self, worksheet):
        title = ['核算维度编码', '核算维度名称', '借方', '贷方', '未回票金额', '合计欠款金额']
        worksheet.append(title)
        sortedList = sorted(self.result, key=lambda x:x[0])
        number_format_style = NamedStyle(name='number_format_style', number_format=numbers.FORMAT_NUMBER_COMMA_SEPARATED1)

        index = 2
        for item in sortedList:
            worksheet.append(item)
            worksheet.cell(index, 3).style = number_format_style
            worksheet.cell(index, 4).style = number_format_style
            worksheet.cell(index, 5).style = number_format_style
            worksheet.cell(index, 6).style = number_format_style
            index = index + 1


dialog = honmeDialog()
dialog.show()
