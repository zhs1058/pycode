import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side, numbers
# 原始表和最终表银行名称对照字典
bank_name_map = {"东城区支行": "东城", "西城区支行": "西城", "金融大街支行": "金融街", "朝阳区支行": "朝阳",
                 "望京支行": "望京", "海淀区支行": "海淀", "中关村支行": "中关村", "丰台区支行": "丰台",
                 "石景山区支行": "石景山", "大兴区支行": "大兴", "通州区支行": "通州", "房山区支行": "房山",
                 "顺义区支行": "顺义", "门头沟区支行": "门头沟", "密云区支行": "密云", "延庆区支行": "延庆",
                 "平谷区支行": "平谷", "昌平区支行": "昌平", "怀柔区支行": "怀柔", "亦庄支行": "亦庄"}

total_now = 0.0
total_pre = 0.0
other_now = 0.0
other_pre = 0.0
# 读取项目所在的行号
def read_project_row_num(sheet, target_project):
    target_num = 0
    column_number = 1
    for cell in sheet.iter_cols(min_col=column_number, max_col=column_number, values_only=True):
        for row_number, value in enumerate(cell, start=1):
            if target_project in str(value):
                target_num = row_number
                return target_num
    return target_num


# 读取人员表所在的行号和列号并获取值
def read_person_row_col(sheet, row_name, col_name):
    row_num = 0
    col_num = 0
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # 遍历当前行的每一列
        for col_number, cell_value in enumerate(row, start=1):
            # 检查单元格的值是否包含目标字符串
            if col_name in str(cell_value):
                col_num = col_number
                break
            if row_name in str(cell_value):
                row_num = row_number
                break
    cell_value = sheet.cell(row=row_num, column=col_num).value
    if cell_value is None:
        cell_value = 0
    return cell_value


# 读取支行所在列号和行号
def read_excel(sheet, target_bank_name):
    bank_row_num = 0
    bank_col_num = 0
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # 遍历当前行的每一列
        for col_number, cell_value in enumerate(row, start=1):
            # 检查单元格的值是否包含目标字符串
            if target_bank_name in str(cell_value):
                bank_row_num = row_number
                bank_col_num = col_number
                break
    return bank_row_num, bank_col_num + 1


# 获取某支行某一个项目的本年累计数
def get_amount_bank_item(sheet, row_num, col_num):
    cell_value = sheet.cell(row=row_num, column=col_num).value
    return cell_value


# 获取单元格值
def get_cell_value(sheet, item_name, bank_name):
    item_row_num = read_project_row_num(sheet, item_name)
    bank_row_num, bank_col_num = read_excel(sheet, bank_name)

    return get_amount_bank_item(sheet, item_row_num, bank_col_num)


# 获取sheet页
def get_sheet(wb, sheet_name_target):
    try:
        # 获取所有 sheet 页的名称
        sheet_names = wb.sheetnames

        # 遍历 sheet 页名称，查找包含指定字符串的 sheet 页
        for sheet_name in sheet_names:
            if sheet_name_target in sheet_name:
                return wb[sheet_name]
        return None
    finally:
        print("选择Sheet页结束")
        # 关闭Excel文件
        # wb.close()


# 计算业务及管理费
def get_item_1(sheet_now, sheet_pre, bank_key):
    global total_now
    global  total_pre
    # 获取业务及管理费
    business_expense_now = float(get_cell_value(sheet_now, "业务及管理费", bank_key))

    business_expense_pre = float(get_cell_value(sheet_pre, "业务及管理费", bank_key))

    # 获取储蓄存款代理费
    store_expense_now = float(get_cell_value(sheet_now, "储蓄存款代理费", bank_key))
    store_expense_pre = float(get_cell_value(sheet_pre, "储蓄存款代理费", bank_key))
    # 获取外币存款代理费
    foreign_store_expense_now = float(get_cell_value(sheet_now, "外币存款代理费", bank_key))
    foreign_store_expense_pre = float(get_cell_value(sheet_pre, "外币存款代理费", bank_key))
    # 当期
    business_expense_now = (business_expense_now - store_expense_now - foreign_store_expense_now) / 10000
    total_now = business_expense_now
    # 去年同期
    business_expense_pre = (business_expense_pre - store_expense_pre - foreign_store_expense_pre) / 10000
    total_pre = business_expense_pre
    # 同比增加
    increase = business_expense_now - business_expense_pre
    # 同比增幅
    increase_ratio = 0
    if business_expense_pre != 0:
        increase_ratio = (increase / business_expense_pre) * 100

    result = [round(business_expense_now, 2), round(business_expense_pre, 2), round(increase, 2), str(round(increase_ratio, 2)) + '%']

    return result

# 计算职工工资
def get_item_2(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 获取职工工资
    # 当期
    wage_now = float(get_cell_value(sheet_now, "职工工资", bank_key)) / 10000
    other_now += wage_now
    # 去年同期
    wage_pre = float(get_cell_value(sheet_pre, "职工工资", bank_key)) / 10000
    other_pre += wage_pre
    # 同比增加
    increase = wage_now - wage_pre
    # 同比增幅
    increase_ratio = 0
    if wage_pre != 0:
        increase_ratio = (increase / wage_pre) * 100

    result = [round(wage_now, 2), round(wage_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '', '']

    return result

# 正式工人数
def get_item_3(sheet_now, sheet_pre, bank_key):
    r1_now_1 = int(read_person_row_col(sheet_now, bank_key, "合同用工"))
    r1_now_2 = int(read_person_row_col(sheet_now, bank_key, "保留关系"))
    r1_pre = int(read_person_row_col(sheet_pre, bank_key, "合同用工"))
    r1_now = r1_now_1 + r1_now_2
    # 同比增加
    increase = r1_now - r1_pre

    result = [r1_now, r1_pre, increase]

    return result

# 劳务工人数
def get_item_4(sheet_now, sheet_pre, bank_key):
    r1_now = int(read_person_row_col(sheet_now, bank_key, "劳务用工"))
    r1_pre = int(read_person_row_col(sheet_pre, bank_key, "劳务用工"))

    # 同比增加
    increase = r1_now - r1_pre

    result = [r1_now, r1_pre, increase]

    return result

# 劳务性支出
def get_item_5(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    labour_now = float(get_cell_value(sheet_now, "劳务性支出", bank_key)) / 10000
    other_now += labour_now
    # 去年同期
    labour_pre = float(get_cell_value(sheet_pre, "劳务性支出", bank_key)) / 10000
    other_pre += labour_pre
    # 同比增加
    increase = labour_now - labour_pre
    # 同比增幅
    increase_ratio = 0
    if labour_pre != 0:
        increase_ratio = (increase / labour_pre) * 100

    result = [round(labour_now, 2), round(labour_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%']

    return result

# 社保、住房公积金
def get_item_6(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r1_now = float(get_cell_value(sheet_now, "社会保险费", bank_key)) / 10000
    r2_now = float(get_cell_value(sheet_now, "补充养老保险费", bank_key)) / 10000
    r3_now = float(get_cell_value(sheet_now, "补充医疗保险费", bank_key)) / 10000
    r4_now = float(get_cell_value(sheet_now, "住房公积金", bank_key)) / 10000
    # 去年同期
    r1_pre = float(get_cell_value(sheet_pre, "社会保险费", bank_key)) / 10000
    r2_pre = float(get_cell_value(sheet_pre, "补充养老保险费", bank_key)) / 10000
    r3_pre = float(get_cell_value(sheet_pre, "补充医疗保险费", bank_key)) / 10000
    r4_pre = float(get_cell_value(sheet_pre, "住房公积金", bank_key)) / 10000

    r_now = r1_now + r2_now + r3_now + r4_now
    r_pre = r1_pre + r2_pre + r3_pre + r4_pre
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%']

    return result


# 工会经费+教育经费
def get_item_7(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r1_now = float(get_cell_value(sheet_now, "工会经费", bank_key)) / 10000
    r2_now = float(get_cell_value(sheet_now, "职工教育经费", bank_key)) / 10000
    # 去年同期
    r1_pre = float(get_cell_value(sheet_pre, "工会经费", bank_key)) / 10000
    r2_pre = float(get_cell_value(sheet_pre, "职工教育经费", bank_key)) / 10000

    r_now = r1_now + r2_now
    r_pre = r1_pre + r2_pre
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%']

    return result


# 福利费
def get_item_8(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "职工福利费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "职工福利费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 宣传费
def get_item_9(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "业务宣传费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "业务宣传费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 房租相关费用
def get_item_10(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r1_now = float(get_cell_value(sheet_now, "租赁费", bank_key)) / 10000
    r2_now = float(get_cell_value(sheet_now, "租赁款利息费", bank_key)) / 10000
    r3_now = float(get_cell_value(sheet_now, "使用权资产折旧费", bank_key)) / 10000
    # 去年同期
    r1_pre = float(get_cell_value(sheet_pre, "租赁费", bank_key)) / 10000
    r2_pre = float(get_cell_value(sheet_pre, "租赁款利息费", bank_key)) / 10000
    r3_pre = float(get_cell_value(sheet_pre, "使用权资产折旧费", bank_key)) / 10000

    r_now = r1_now + r2_now + r3_now
    r_pre = r1_pre + r2_pre + r3_pre
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 物业费
def get_item_11(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "物业管理费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "物业管理费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 水电取暖费
def get_item_12(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "水电取暖费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "水电取暖费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 钞币运送费
def get_item_13(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "钞币运送费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "钞币运送费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 代理公司存款营销费
def get_item_14(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "公司存款营销费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "公司存款营销费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100


    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 折旧费及摊销
def get_item_15(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r1_now = float(get_cell_value(sheet_now, "固定资产折旧费", bank_key)) / 10000
    r2_now = float(get_cell_value(sheet_now, "无形资产摊销", bank_key)) / 10000
    r3_now = float(get_cell_value(sheet_now, "长期待摊费用摊销", bank_key)) / 10000
    # 去年同期
    r1_pre = float(get_cell_value(sheet_pre, "固定资产折旧费", bank_key)) / 10000
    r2_pre = float(get_cell_value(sheet_pre, "无形资产摊销", bank_key)) / 10000
    r3_pre = float(get_cell_value(sheet_pre, "长期待摊费用摊销", bank_key)) / 10000

    r_now = r1_now + r2_now + r3_now
    r_pre = r1_pre + r2_pre + r3_pre
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 外包费
def get_item_16(sheet_now, sheet_pre, bank_key):
    global other_now
    global other_pre
    # 当期
    r_now = float(get_cell_value(sheet_now, "外包费", bank_key)) / 10000
    # 去年同期
    r_pre = float(get_cell_value(sheet_pre, "外包费", bank_key)) / 10000
    other_now += r_now
    other_pre += r_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']

    return result


# 其他项目
def get_item_17(sheet_now, sheet_pre, bank_key):
    global other_now, other_pre, total_now, total_pre


    r_now = total_now - other_now
    r_pre = total_pre - other_pre
    # 同比增加
    increase = r_now - r_pre
    # 同比增幅
    increase_ratio = 0
    if r_pre != 0:
        increase_ratio = (increase / r_pre) * 100

    result = [round(r_now, 2), round(r_pre, 2), round(increase, 2),
              str(round(increase_ratio, 2)) + '%', '']
    return result


# 正式工人数合计
def get_item_3_total(sheet_now, sheet_pre, bank_key):
    r1_now_1 = int(read_person_row_col(sheet_now, "合计", "合同用工"))
    r1_now_2 = int(read_person_row_col(sheet_now, "合计", "保留关系"))

    r1_pre_1 = int(read_person_row_col(sheet_pre, "合计", "合同用工"))
    r1_pre_2 = int(read_person_row_col(sheet_pre, "合计", "保留关系"))
    r1_now = r1_now_1 + r1_now_2
    r1_pre = r1_pre_1 + r1_pre_2
    # 同比增加
    increase = r1_now - r1_pre

    result = [r1_now, r1_pre, increase]

    return result



# 劳务工人数合计
def get_item_4_total(sheet_now, sheet_pre, bank_key):
    r1_now = int(read_person_row_col(sheet_now, "合计", "劳务用工"))

    r1_pre = int(read_person_row_col(sheet_pre, "合计", "劳务用工"))

    # 同比增加
    increase = r1_now - r1_pre

    result = [r1_now, r1_pre, increase]

    return result


# 正式工人数
def get_item_3_branch(sheet_now, sheet_pre, bank_key):
    r1_now_1 = int(read_person_row_col(sheet_now, "分行机关", "合同用工"))
    r1_now_2 = int(read_person_row_col(sheet_now, "分行机关", "保留关系"))
    r2_now_1 = int(read_person_row_col(sheet_now, "运营中心", "合同用工"))
    r2_now_2 = int(read_person_row_col(sheet_now, "运营中心", "保留关系"))
    r3_now_1 = int(read_person_row_col(sheet_now, "分行营业部", "合同用工"))
    r3_now_2 = int(read_person_row_col(sheet_now, "分行营业部", "保留关系"))

    r1_pre_1 = int(read_person_row_col(sheet_pre, "分行机关", "合同用工"))
    r1_pre_2 = int(read_person_row_col(sheet_pre, "分行机关", "保留关系"))
    r2_pre_1 = int(read_person_row_col(sheet_pre, "运营中心", "合同用工"))
    r2_pre_2 = int(read_person_row_col(sheet_pre, "运营中心", "保留关系"))
    r3_pre_1 = int(read_person_row_col(sheet_pre, "分行营业部", "合同用工"))
    r3_pre_2 = int(read_person_row_col(sheet_pre, "分行营业部", "保留关系"))
    r1_now = r1_now_1 + r1_now_2 + r2_now_1 + r2_now_2 + r3_now_1 + r3_now_2
    r1_pre = r1_pre_1 + r1_pre_2 + r2_pre_1 + r2_pre_2 + r3_pre_1 + r3_pre_2
    # 同比增加
    increase = r1_now - r1_pre

    result = [r1_now, r1_pre, increase]

    return result

# 劳务工人数
def get_item_4_branch(sheet_now, sheet_pre, bank_key):
    r1_now_1 = int(read_person_row_col(sheet_now, "分行机关", "劳务用工"))
    r1_now_2 = int(read_person_row_col(sheet_now, "运营中心", "劳务用工"))
    r1_now_3 = int(read_person_row_col(sheet_now, "分行营业部", "劳务用工"))
    r1_now = r1_now_1 + r1_now_2 + r1_now_3

    r1_pre_1 = int(read_person_row_col(sheet_pre, "分行机关", "劳务用工"))
    r1_pre_2 = int(read_person_row_col(sheet_pre, "运营中心", "劳务用工"))
    r1_pre_3 = int(read_person_row_col(sheet_pre, "分行营业部", "劳务用工"))
    r1_pre = r1_pre_1 + r1_pre_2 + r1_pre_3

    # 同比增加
    increase = r1_now - r1_pre

    result = [r1_now, r1_pre, increase]

    return result

# 计算小计
def count_total_1(list1, list2):
    result_item = []
    result_item.append("小计")
    list_len_1 = len(list1)
    list_len_2 = len(list2)
    if(list_len_1 == list_len_2):
        for i in range(1, list_len_1):
            item1 = list1[i]
            item2 = list2[i]
            if type(item1) == float and type(item2) == float:
                result_item.append(round((item2 - item1), 2))
            elif type(item1) == int and type(item2) == int:
                result_item.append(item2 - item1)
            elif type(item1) == str and type(item2) == str and '%' in item1 and '%' in item2:
                increase = result_item[i-1]
                pre = result_item[i-2]
                increase_ratio = 0
                if pre != 0:
                    increase_ratio = (increase / pre) * 100
                result_item.append(str(round(increase_ratio, 2)) + '%')
            else:
                result_item.append('')



    return result_item





# 计算分行
def count_total_2(sheet_now, sheet_pre, sheet_person_now, sheet_person_pre):
    global other_now, other_pre, total_now, total_pre

    result_item = []
    total_now = 0.0
    total_pre = 0.0
    other_now = 0.0
    other_pre = 0.0
    key = "11000013"
    result_item.append("分行")
    result_item += get_item_1(sheet_now, sheet_pre, key)
    result_item += get_item_2(sheet_now, sheet_pre, key)
    result_item += get_item_3_branch(sheet_person_now, sheet_person_pre, key)
    result_item += get_item_4_branch(sheet_person_now, sheet_person_pre, key)
    result_item += get_item_5(sheet_now, sheet_pre, key)
    result_item += get_item_6(sheet_now, sheet_pre, key)
    result_item += get_item_7(sheet_now, sheet_pre, key)
    result_item += get_item_8(sheet_now, sheet_pre, key)
    result_item += get_item_9(sheet_now, sheet_pre, key)
    result_item += get_item_10(sheet_now, sheet_pre, key)
    result_item += get_item_11(sheet_now, sheet_pre, key)
    result_item += get_item_12(sheet_now, sheet_pre, key)
    result_item += get_item_13(sheet_now, sheet_pre, key)
    result_item += get_item_14(sheet_now, sheet_pre, key)
    result_item += get_item_15(sheet_now, sheet_pre, key)
    result_item += get_item_16(sheet_now, sheet_pre, key)
    result_item += get_item_17(sheet_now, sheet_pre, key)

    return result_item

# 计算合计
def count_total_3(sheet_now, sheet_pre, sheet_person_now, sheet_person_pre):
    global other_now, other_pre, total_now, total_pre

    result_item = []
    total_now = 0.0
    total_pre = 0.0
    other_now = 0.0
    other_pre = 0.0
    key = "北京分行（合并）"
    result_item.append("合计")
    result_item += get_item_1(sheet_now, sheet_pre, key)
    result_item += get_item_2(sheet_now, sheet_pre, key)
    result_item += get_item_3_total(sheet_person_now, sheet_person_pre, key)
    result_item += get_item_4_total(sheet_person_now, sheet_person_pre, key)
    result_item += get_item_5(sheet_now, sheet_pre, key)
    result_item += get_item_6(sheet_now, sheet_pre, key)
    result_item += get_item_7(sheet_now, sheet_pre, key)
    result_item += get_item_8(sheet_now, sheet_pre, key)
    result_item += get_item_9(sheet_now, sheet_pre, key)
    result_item += get_item_10(sheet_now, sheet_pre, key)
    result_item += get_item_11(sheet_now, sheet_pre, key)
    result_item += get_item_12(sheet_now, sheet_pre, key)
    result_item += get_item_13(sheet_now, sheet_pre, key)
    result_item += get_item_14(sheet_now, sheet_pre, key)
    result_item += get_item_15(sheet_now, sheet_pre, key)
    result_item += get_item_16(sheet_now, sheet_pre, key)
    result_item += get_item_17(sheet_now, sheet_pre, key)

    return result_item


# 设置列宽
def set_excel_col_width(worksheet):
    for i in range(1,80):
        if i == 39 or i == 44 or i ==53:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 36
        elif i == 79:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 54
        else:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18





# 设置表头格式
def set_excel_head_1_style(worksheet):
    # 表头第一行合并范围
    merge_row_1 = [(1, 1, 79, 1)]
    start_col = 1
    start_row = 1
    end_col = 79
    end_row = 1
    merge_range_str = f'{openpyxl.utils.get_column_letter(start_col)}{start_row}:{openpyxl.utils.get_column_letter(end_col)}{end_row}'
    worksheet.merge_cells(merge_range_str)
    # 设置第一行样式
    font_row_1 = Font(name="宋体", size=20)
    alignment_row_1 = Alignment(vertical="center", horizontal="center", wrap_text=True)
    worksheet.cell(1, 1).font = font_row_1
    worksheet.cell(1, 1).alignment = alignment_row_1
    worksheet.cell(1, 1, '成本支出情况表')


# 设置表头第二行
def set_excel_head_2_style(worksheet):
    font_row_1 = Font(name="宋体", size=12)
    alignment_row_1 = Alignment(vertical="center", horizontal="left", wrap_text=True)
    worksheet.cell(2, 78).font = font_row_1
    worksheet.cell(2, 78).alignment = alignment_row_1
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(78)].width = 18
    worksheet.cell(2, 78, '单位：万元')


# 设置表头第四行格式
def set_excel_head_4_style(worksheet):
    font_row = Font(name="宋体", size=12)
    alignment_row = Alignment(vertical="center", horizontal="center", wrap_text=True)
    # fill = PatternFill(fill_type="solid", fgColor="808080")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                    bottom=Side(style="thin"))
    # 表头合并范围
    # (start_col, start_row, end_col, end_row)

    for i in range(1,80):
        worksheet.cell(4, i).border = border

    value_dict = {2: "业务及管理费", 6: "职工工资", 12: "正式工人数", 15: "劳务工人数", 18: "劳务性支出", 22: "社保、住房公积金", 26: "工会经费+教育经费", 30: "福利费", 35: "宣传费", 40: "房租相关费用", 45: "物业费", 50: "水电取暖费", 55: "钞币运送费", 60: "代理公司存款营销费", 65: "折旧费及摊销", 70: "外包费", 75: "其他项目"}
    merge_row_list = [(2, 4, 5, 4), (6, 4, 11, 4), (12, 4, 14, 4), (15, 4, 17, 4), (18, 4, 21, 4), (22, 4, 25, 4), (26, 4, 29, 4), (30, 4, 34, 4), (35, 4, 39, 4), (40, 4, 44, 4), (45, 4, 49, 4), (50, 4, 54, 4), (55, 4, 59, 4), (60, 4, 64, 4), (65, 4, 69, 4), (70, 4, 74, 4), (75, 4, 78, 4)]
    for merge_range in merge_row_list:
        start_col, start_row, end_col, end_row = merge_range
        merge_range_str = f'{openpyxl.utils.get_column_letter(start_col)}{start_row}:{openpyxl.utils.get_column_letter(end_col)}{end_row}'
        worksheet.merge_cells(merge_range_str)
        worksheet.cell(4, start_col).font = font_row
        worksheet.cell(4, start_col).alignment = alignment_row
        worksheet.cell(4, start_col, value_dict.get(start_col))

    worksheet.cell(4, 1).font = font_row
    worksheet.cell(4, 1).alignment = alignment_row
    worksheet.cell(4, 1, "项目")
    worksheet.cell(4, 79).font = font_row
    worksheet.cell(4, 79).alignment = alignment_row
    worksheet.cell(4, 79, "备注其他情况")

    worksheet.row_dimensions[4].height = 30


#  设置第五行格式
def set_excel_head_5_style(worksheet):
    font_row = Font(name="宋体", size=12)
    alignment_row = Alignment(vertical="center", horizontal="center", wrap_text=True)
    # fill = PatternFill(fill_type="solid", fgColor="808080")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                    bottom=Side(style="thin"))
    col_1 = "202311"
    col_2 = "202211"
    col_3 = "同比增加"
    col_4 = "同比增幅"
    data_to_insert = ['单位', col_1, col_2, col_3, col_4, col_1, col_2, col_3, col_4, '绩效考核结果', '备注原因', col_1,
                      col_2, col_3, col_1, col_2, col_3,
                      col_1, col_2, col_3, col_4, col_1, col_2, col_3, col_4, col_1, col_2, col_3, col_4, col_1, col_2,
                      col_3, col_4, '备注',
                      col_1, col_2, col_3, col_4, '备注', col_1, col_2, col_3, col_4, '备注', col_1, col_2, col_3,
                      col_4, '备注',
                      col_1, col_2, col_3, col_4, '备注', col_1, col_2, col_3, col_4, '备注', col_1, col_2, col_3,
                      col_4, '备注',
                      col_1, col_2, col_3, col_4, '备注', col_1, col_2, col_3, col_4, '备注', col_1, col_2, col_3,
                      col_4, ]
    worksheet.append(data_to_insert)
    # 表头合并范围
    # (start_col, start_row, end_col, end_row)

    for i in range(1, 80):
        worksheet.cell(5, i).font = font_row
        worksheet.cell(5, i).alignment = alignment_row
        worksheet.cell(5, i).border = border

    worksheet.row_dimensions[5].height = 30
    merge_range_str = f'{openpyxl.utils.get_column_letter(79)}{4}:{openpyxl.utils.get_column_letter(79)}{5}'
    worksheet.merge_cells(merge_range_str)



def set_excel_data(worksheet, data):
    dataLength = len(data)
    print('我的长度',dataLength)
    font_row = Font(name="宋体", size=12)
    alignment_row = Alignment(vertical="center", horizontal="center", wrap_text=True)
    # fill = PatternFill(fill_type="solid", fgColor="808080")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                    bottom=Side(style="thin"))
    for item in data:
        worksheet.append(item)

    number_format_style = NamedStyle(name='number_format_style', number_format=numbers.FORMAT_NUMBER_COMMA_SEPARATED1)
    percent_format_style = NamedStyle(name='percent_format_style', number_format=numbers.FORMAT_PERCENTAGE_00)
    percent_list = [5, 9, 21, 25, 29, 33, 38, 43, 48, 53, 58, 63, 68, 73, 78]
    number_list = [12, 13, 14, 15, 16, 17]
    for i in range(6, dataLength + 6):
        worksheet.row_dimensions[i].height = 90
        for j in range(1, 80):
            if j in percent_list:
                worksheet.cell(i, j).style = percent_format_style
            elif j not in number_list:
                worksheet.cell(i, j).style = number_format_style
            worksheet.cell(i, j).font = font_row
            worksheet.cell(i, j).alignment = alignment_row
            worksheet.cell(i, j).border = border



# 设置excel表格式
def set_excel_style(output_file_path, data):
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("支出明细表", index=0)

    set_excel_col_width(worksheet)
    set_excel_head_1_style(worksheet)
    set_excel_head_2_style(worksheet)
    set_excel_head_4_style(worksheet)
    set_excel_head_5_style(worksheet)
    set_excel_data(worksheet, data)


    workbook.save(output_file_path)



if __name__ == "__main__":
    # 输入文件路径（支出明细表）
    input_file_path_now = "D:/RPATestDocumet/支出明细表-原表.xlsx"
    input_file_path_pre = "D:/RPATestDocumet/月报明细报表（202211CNY元）.xlsx"

    #人员表路径
    input_file_path_person_now = "D:/RPATestDocumet/person/2023年劳产率每月（11月）（人力原表）.xlsx"
    input_file_path_person_pre = "D:/RPATestDocumet/person/2022年劳产率每月（11月）.xlsx"

    # 打开Excel文件
    wb_now = openpyxl.load_workbook(input_file_path_now, data_only=True)
    wb_pre = openpyxl.load_workbook(input_file_path_pre, data_only=True)

    person_wb_now = openpyxl.load_workbook(input_file_path_person_now, data_only=True)
    person_wb_pre = openpyxl.load_workbook(input_file_path_person_pre, data_only=True)


    # 选择第一个工作表
    sheet_now = get_sheet(wb_now, "支出明细表")
    sheet_pre = get_sheet(wb_pre, "支出明细表")

    # 选择最后一个工作表
    sheet_person_now = person_wb_now.worksheets[-1]
    sheet_person_pre = person_wb_pre.worksheets[-1]


    # 输出文件路径（新的Excel表格）
    output_file_path = "D:/RPATestDocumet/text.xlsx"

    result = []
    for key, value in bank_name_map.items():
        result_item = []
        result_item.append(value)
        result_item += get_item_1(sheet_now, sheet_pre, key)
        result_item += get_item_2(sheet_now, sheet_pre, key)
        result_item += get_item_3(sheet_person_now, sheet_person_pre, key)
        result_item += get_item_4(sheet_person_now, sheet_person_pre, key)
        result_item += get_item_5(sheet_now, sheet_pre, key)
        result_item += get_item_6(sheet_now, sheet_pre, key)
        result_item += get_item_7(sheet_now, sheet_pre, key)
        result_item += get_item_8(sheet_now, sheet_pre, key)
        result_item += get_item_9(sheet_now, sheet_pre, key)
        result_item += get_item_10(sheet_now, sheet_pre, key)
        result_item += get_item_11(sheet_now, sheet_pre, key)
        result_item += get_item_12(sheet_now, sheet_pre, key)
        result_item += get_item_13(sheet_now, sheet_pre, key)
        result_item += get_item_14(sheet_now, sheet_pre, key)
        result_item += get_item_15(sheet_now, sheet_pre, key)
        result_item += get_item_16(sheet_now, sheet_pre, key)
        result_item += get_item_17(sheet_now, sheet_pre, key)
        total_now = 0.0
        total_pre = 0.0
        other_now = 0.0
        other_pre = 0.0

        result.append(result_item)
    list2 = count_total_2(sheet_now, sheet_pre, sheet_person_now, sheet_person_pre)
    list3 = count_total_3(sheet_now, sheet_pre, sheet_person_now, sheet_person_pre)
    list1 = count_total_1(list2, list3)

    result.append(list1)
    result.append(list2)
    result.append(list3)

    set_excel_style(output_file_path, result)

    # 关闭Excel文件
    wb_now.close()
    wb_pre.close()
    person_wb_now.close()
    person_wb_pre.close()
