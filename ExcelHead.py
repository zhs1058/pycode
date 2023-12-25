import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 创建工作簿和新增Sheet页
filePath = "C:/Users/89704/OneDrive/Documents/我的文档/工作/RPA测试文档/test.xlsx"
workbook = openpyxl.load_workbook(filePath)
worksheet = workbook.create_sheet("新增Sheet3")

# 插入一行，设置列名
worksheet.append(["字段序号", "工程名称（必填）", "表英文名（必填）", "表中文名（必填）", "字段英文名（必填）", "字段中文名（必填）", "字段类型（必填）", "长度（必填）", "精度", "是否主键（必填）", "外键", "是否可以为空（Y/N）(必填）", "缺省值", "取值范围", "业务取值说明", "当前状态（必填）", "字段含义（必填）", "是否贯标（必填）", "企业级数据字典数据项编号（必填）", "不贯标原因（不贯标时必填）", "数据质量规则编号", "数据安全一级分类", "数据安全二级分类", "数据安全三级分类", "数据安全四级分类", "数据安全级别", "数据主管方", "数据生产方", "备注"])

# 设置列头样式
font = Font(name="宋体", size=12, bold=True)
alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
fill = PatternFill(fill_type="solid", fgColor="808080")
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 设置列头单元格样式
for cell in worksheet[1]:
    cell.font = font
    cell.alignment = alignment
    cell.fill = fill
    cell.border = border

# 设置行高
worksheet.row_dimensions[1].height = 76

# 设置列宽
worksheet.column_dimensions["A"].width = 16
worksheet.column_dimensions["B"].width = 16
# 保存Excel文件
workbook.save(filePath)
